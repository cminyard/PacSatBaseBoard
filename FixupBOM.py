#!/usr/bin/python3

# This program read a Kicad BOM CSV output from stdin or from a file
# supplied on the command line and generates a BOM in XLS format
# suitable for JLCPCB on standard output.  This involves rearranging
# the columns and changing the column names and renaming some
# footprints.

import sys
import csv
from openpyxl import Workbook

infn = None
outfn = None
do_stage2_xlat = False
use_murata = False
value_in_comment = False

unknown_components = []

in_flags = True
for i in sys.argv[1:]:
    if in_flags and i.startswith('-'):
        if i == '--':
            in_flags = False
        elif i == '-2':
            do_stage2_xlat = True
        elif i == '-murata':
            use_murata = True
        elif i == '-value-in-comment':
            value_in_comment = True
        else:
            sys.stderr.write("Unknown flag: " + i)
            sys.exit(1)
            pass
        pass
    else:
        in_flags = False
        if infn is None:
            infn = i
        elif outfn is None:
            outfn = i
        else:
            sys.stderr.write("Too many filenames given\n")
            sys.exit(1)
            pass
        pass
    pass

if outfn is None:
    sys.stderr.write("No CSV BOM file and XLS/CSV output file given\n")
    sys.exit(1)
    pass

if not infn.endswith(".csv"):
    sys.stderr.write("First file doesn't end in '.csv': " + infn[1] + "\n")
    sys.exit(1)
    pass

do_xls = True
if outfn.endswith(".csv"):
    do_xls = False
elif not outfn.endswith(".xls") and not outfn.endswith(".xlsx"):
    sys.stderr.write("Second file doesn't end in '.xls' or '.xlsx' or '.csv': "
                     + outfn + "\n")
    sys.exit(1)

f = open(infn)

cf = csv.reader(f, delimiter=';')
line = cf.__next__()
if len(line) != 7:
    sys.stderr.write("First line doesn't have 7 values, doesn't appear to be"
                     + " a Kicad BOM output")
    sys.exit(1)
    pass

expected_first_line = 'Id;Designator;Footprint;Quantity;Designation;Supplier and ref;'.split(";")
for i in range(0, len(expected_first_line)):
    if line[i] != expected_first_line[i]:
        sys.stderr.write("First line pos %d: Expected %s, got %s" %
                         (i, expected_first_line[i], line[i]))
        sys.exit(1)
        pass
    pass

footprint_xlats = {
    'R_0402_1005Metric': '0402',
    'R_0603_1608Metric': '0603',
    'R_0805_2012Metric': '0805',
    'R_1206_3216Metric': '1206',
    'C_0402_1005Metric': '0402',
    'C_0603_1608Metric': '0603',
    'C_0805_2012Metric': '0805',
    'C_1206_3216Metric': '1206',
    'L_0402_1005Metric': '0402',
    'L_0603_1608Metric': '0603',
    'L_0805_2012Metric': '0805',
    'L_1206_3216Metric': '1206',
    'D_0603_1608Metric': '0603',
    'D_SOD_323':         'SOD-323',
    'D_SOD_882':         'SOD-882',

    'C_0603_1608Metric_Pad1.08x0.95mm_HandSolder': '0603',
    'C_0805_2012Metric_Pad1.18x1.45mm_HandSolder': '0805',
    'R_0603_1608Metric_Pad0.98x0.95mm_HandSolder': '0603',
    'R_1206_3216Metric_Pad1.30x1.75mm_HandSolder': '1206',
    'L_0603_1608Metric_Pad1.05x0.95mm_HandSolder': '0603',
    'C_0402_1005Metric_Pad0.74x0.62mm_HandSolder': '0402',
    'LED_0603_1608Metric_Pad1.05x0.95mm_HandSolder': '0603',
    'C_1206_3216Metric_Pad1.33x1.80mm_HandSolder': '1206',
    'L_1210_3225Metric_Pad1.42x2.65mm_HandSolder': '1210',
    'L_0402_1005Metric_Pad0.77x0.64mm_HandSolder': '0402',
    'R_0402_1005Metric_Pad0.72x0.64mm_HandSolder': '0402',
    'L_0805_2012Metric_Pad1.05x1.20mm_HandSolder': '0805',

    'SOT-23-5_HandSoldering': 'SOT-23-5',
    'SOT-23-6_HandSoldering': 'SOT-23-6',
}
def xlat_footprint(s):
    if s in footprint_xlats:
        return footprint_xlats[s]
    return s

# Mostly Coilcraft parts.  All are automotive certified and have
# somewhat higher Q values than the Murata parts.
value_to_partnum_xlats_1 = {
    ('1nH',     	'0402'): ('Coilcraft',	'0402DC-1N0XJRW '),
    ('3.0nH 2%',	'0402'): ('Coilcraft',	'0402DC-3N0XGRW'),
    ('5.8nH 2%',	'0402'): ('Coilcraft',	'0402DC-5N8XGRW'),
    ('6.0nH 2%',	'0402'): ('Coilcraft',	'0402DC-6N0XGRW'),
    ('11nH 2%', 	'0603'): ('Coilcraft',	'0603DC-11NXGRW'),
    ('16nH 2%', 	'0402'): ('Coilcraft',	'0402DC-16NXGRW'),
    ('17nH 2%', 	'0402'): ('Murata',	'LQW15AN17NG8Z'),
    ('18nH 2%', 	'0603'): ('Coilcraft',	'0603DC-18NXGRW'),
    ('18nH 2%', 	'0402'): ('Coilcraft',	'0402DC-18NXGRW'),
    ('20nH 2%', 	'0402'): ('Coilcraft',	'0402DC-20NXGRW'),
    ('22nH 2%', 	'0402'): ('Coilcraft',	'0402DC-22NXGRW'),
    ('27nH 2%', 	'0603'): ('Coilcraft',	'0603DC-27NXGRW'),
    ('36nH I>1A',	'0603'): ('Coilcraft',	'0603DC-36NXGRW'),
    ('43nH 2%', 	'0603'): ('Coilcraft',	'0603DC-43NXGRW'),
    ('43nH 2%', 	'0805'): ('Coilcraft',	'0805CS-430XGRC'),
    ('47nH 2%', 	'0603'): ('Coilcraft',	'0603DC-47NXGRW'),
    ('47nH 2%', 	'0805'): ('Coilcraft',	'0805CS-470XGRC'),
    ('68nH',    	'0603'): ('Coilcraft',	'0603DC-68NXJRW'),
    ('78nH 2%', 	'0603'): ('Murata',	'LQW18AN78NG8ZD'),
    ('91nH 2%', 	'0805'): ('Coilcraft',	'0805CS-910XGRC'),
    ('100nH',   	'0603'): ('Murata',	'LQW18CNR10K0ZD'),
    ('100nH I>1A',   	'0603'): ('Murata',	'LQW18CNR10K0ZD'),
    ('180nH 2%',	'0805'): ('Coilcraft',	'0805CS-181XGRC'),
    ('470nH 2%',        '0603'): ('Murata',     'LQW18ANR47G0ZD'),
    ('470nH',   	'0805'): ('Coilcraft',	'0805CS-471XGRC'),
    ('1uH',   'L_Murata_DFE201610P'): ('Murata',	'DFE201612PD-1R0M'),
    ('3.3uH', '1210'):                ('Murata',	'DFE322520FD-3R3M'),
    ('',	''): ('',	''),
}

# These are all Murata parts.  The 0805 parts are not automotive
# certified.
value_to_partnum_xlats_1b = {
    ('1nH',     	'0402'): ('Murata',	'LQG15WZ1N0B02D'),
    ('3.0nH 2%',	'0402'): ('Murata',	'LQW15AS3N0G8ZD'),
    ('5.8nH 2%',	'0402'): ('Murata',	'LQW15AN5N8G8ZD'),
    ('6.0nH 2%',	'0402'): ('Murata',	'LQW15AN6N0G8ZD'),
    ('11nH 2%', 	'0603'): ('Murata',	'LQW18AS11NG0ZD'),
    ('17nH 2%', 	'0402'): ('Murata',	'LQW15AN17NG8Z'),
    ('18nH 2%', 	'0603'): ('Murata',	'LQW18AS18NG0ZD'),
    ('18nH 2%', 	'0402'): ('Murata',	'LQW15AN18NG8ZD'),
    ('20nH 2%', 	'0402'): ('Murata',	'LQW15AN20NG8ZD'),
    ('22nH 2%', 	'0402'): ('Murata',	'LQW15AN22NG8ZD'),
    ('27nH 2%', 	'0603'): ('Murata',	'LQW18AS27NG0ZD'),
    ('36nH I>1A',	'0603'): ('Murata',	'LQW18AS36NG0ZD'),
    ('43nH 2%', 	'0603'): ('Murata',	'LQW18AS43NG0ZD'),
    ('43nH 2%', 	'0805'): ('Murata',	'LQW2BAN43NG00L'),
    ('47nH 2%', 	'0603'): ('Murata',	'LQW18AS47NG0ZD'),
    ('47nH 2%', 	'0805'): ('Murata',	'LQW2BAN47NG00L'),
    ('68nH',    	'0603'): ('Murata',	'LQW18AS68NG0ZD'),
    ('78nH 2%', 	'0603'): ('Murata',	'LQW18AN78NG8ZD'),
    ('91nH 2%', 	'0805'): ('Murata',	'LQW2BAN91NG00L'),
    ('100nH',   	'0603'): ('Murata',	'LQW18ASR10G0ZD'),
    ('100nH I>1A',   	'0603'): ('Murata',	'LQW18CNR10K0ZD'),
    ('180nH 2%',	'0805'): ('Murata',	'LQW2BANR18G00L'),
    ('470nH 2%',        '0603'): ('Murata',     'LQW18ANR47G0ZD'),
    ('470nH',   	'0805'): ('Murata',	'LQW21FTR47M0HL'),
    ('1uH',   'L_Murata_DFE201610P'): ('Murata',	'DFE201612PD-1R0M'),
    ('3.3uH', '1210'):                ('Murata',	'DFE322520FD-3R3M'),
    ('4.7uH', '1210'):                ('Murata',	'DFE322520FD-4R7M'),
    ('',	''): ('',	''),
}

# General passive parts.
value_to_partnum_xlats_2 = {
    ('6.8pF',		'0402'): ('Murata',     'GCM1555C1H6R8BA16J'),
    ('10pF',		'0402'): ('Murata',	'GCM1555C1H100JA16D'),
    ('100pF',		'0402'): ('Murata',	'GCM1555C1H101JA16D'),
    ('1nF',		'0402'): ('Murata',	'GCM1555C1H102JA16D'),
    ('4.7nF',		'0402'): ('Kyocera',	'04025C472J4T2A'),
    ('10nF',		'0402'): ('Murata',	'GCM155R71H103KA55D'),
    ('100nF',		'0402'): ('Murata',	'GCM155R71C104KA55D'),
    ('.47uF',           '0603'): ('Murata',     'GCM188R71E474KA64J'),
    ('1uF',		'0603'): ('Murata',	'GCM188R71C105KA64J'),
    ('1uF',		'0805'): ('Murata',	'GCM21BR71E105KA56L'),
    ('4.7uF',		'0805'): ('Murata',	'GCM21BR71C475KA73L'),
    ('10uF',		'0805'): ('Murata',	'GCM21BR71A106KE22K'),
    ('10uF 10V',	'1206'): ('Murata',	'GCM31CR71C106KA64L'),
    ('22uF',		'1206'): ('Murata',	'GCM31CR70J226KE23L'),
    ('47uF 10V',	'1206'): ('Murata',	'GRT31CR61A476KE13L'),
    ('',	''): ('',	''),

    ('0.75pF ±.05 30V',	'0402'): ('Murata',	'GCQ1555C1HR75WB01D'),
    ('1pF ±.05',  	'0402'): ('Murata',	'GCQ1555C1H1R0WB01D'),
    ('2.0pF ±.05 30V',	'0402'): ('Murata',	'GCQ1555C1H2R0WB01D'),
    ('2.2pF ±.05 30V',	'0402'): ('Murata',	'GCQ1555C1H2R2WB01D'),
    ('4.3pF 1%',	'0402'): ('Murata',	'GCQ1555C1H4R3BB01D'),
    ('5.1pF ±.1',	'0402'): ('Murata',	'GCQ1555C1H5R1BB01D'),
    ('5.4pF ±.05 30V',	'0402'): ('Murata',	'GCQ1555C1H5R4BB01D'),
    ('5.6pF ±.1',	'0402'): ('Murata',	'GCQ1555C1H5R6BB01D'),
    ('6.2pF ±.1 30V',	'0402'): ('Murata',	'GCQ1555C1H6R2BB01D'),
    ('6.4pF ±.1 30V',	'0402'): ('Murata',	'GCQ1555C1H6R4BB01D'),
    ('6.5pF ±.1',	'0402'): ('Murata',	'GCQ1555C1H6R5BB01D'),
    ('6.8pF 1% 30V', 	'0402'): ('Murata',	'GCM1555C1H6R8FA16D'),
    ('7.3pF ±.1 30V',	'0402'): ('Murata',	'GCQ1555C1H7R3BB01D'),
    ('11pF 1%', 	'0402'): ('Murata',	'GCQ1555C1H110FB01D'),
    ('11pF 1% 30V', 	'0402'): ('Murata',	'GCQ1555C1H110FB01D'),
    ('12pF 1% 30V', 	'0402'): ('Murata',	'GCQ1555C1H120FB01D'),
    ('12pF 1%', 	'0402'): ('Murata',	'GCQ1555C1H120FB01D'),
    ('15pF 1%', 	'0402'): ('Murata',	'GCQ1555C1H150FB01D'),
    ('18pF 1% 30V', 	'0402'): ('Murata',	'GCQ1555C1H180FB01D'),
    ('20pF 1% 30V', 	'0402'): ('Murata',	'GCQ1555C1H200FB01D'),
    ('22pF 1%', 	'0402'): ('Murata',	'GCQ1555C1H220FB01D'),
    ('27pF 1% 30V', 	'0402'): ('Murata',	'GCM1555C1H270FA16D'),
    ('36pF 1%', 	'0402'): ('Murata',	'GCQ1555C1H360FB01D'),
    ('47pF 1%', 	'0402'): ('Murata',	'GCM1885C2A470FA16D'),
    ('62pF 1%', 	'0402'): ('Murata',	'GCM1555C1H620FA16D'),
    ('68pF 1%', 	'0402'): ('Murata',	'GCM1555C1H680FA16D'),
    ('110pF 1% 30V',	'0402'): ('Murata',	'GCM1555C1H111FA16D'),
    ('',	''): ('',	    ''),

    ('0',       '0402'): ('Panasonic',	'ERJ-2GE0R00X'),
    ('0Ω',	'0402'): ('Panasonic',	'ERJ-2GE0R00X'),
    ('22Ω',	'0402'): ('Panasonic',	'ERA-2AKD220X'),
    ('33Ω',	'0402'): ('Panasonic',	'ERA-2AKD330X'),
    ('50Ω',	'0402'): ('Panasonic',	'ERA-2AED49R9X'),
    ('61.9Ω',	'0402'): ('Panasonic',	'ERA-2AEB61R9X'),
    ('240Ω',	'0402'): ('Panasonic',	'ERA-2AED241X'),
    ('280Ω',	'0402'): ('Panasonic',	'ERA-2AEB2800X'),
    ('470Ω',	'0402'): ('Panasonic',	'ERA-2AED471X'),
    ('1KΩ',	'0402'): ('Panasonic',	'ERA-2AED102X'),
    ('4.7KΩ',	'0402'): ('Panasonic',	'ERA-2AEB472X'),
    ('10KΩ',	'0402'): ('Panasonic',	'ERA-2AED103X'),
    ('18KΩ',	'0402'): ('Panasonic',	'ERA-2AED183X'),
    ('47KΩ',	'0402'): ('Panasonic',	'ERA-2AED473X'),
    ('100KΩ',	'0402'): ('Panasonic',	'ERA-2AED104X'),
    ('1MΩ',	'0402'): ('Panasonic',	'ERJ-U02J105X'),

    ('0',	'0603'): ('Panasonic',	'ERJ-S030R00V'),
    ('0Ω',	'0603'): ('Panasonic',	'ERJ-S030R00V'),
    ('2.4KΩ',	'0603'): ('Panasonic',	'ERJ-PA3D2401V'),
    ('3KΩ',	'0603'): ('Panasonic',	'ERJ-UP3F3001V'),
    ('10KΩ',	'0603'): ('Panasonic',	'ERJ-U03F3001V'),
    ('Ω',	''): ('',    ''),
    
    ('50Ω >=2W',        '1206'): ('Rohm',	'ESR18EZPF49R9'),

    ('25mΩ 1%', 	'0402'): ('Yaego',	'PE0402FRF470R025L'),
    ('3.32KΩ 1%',	'0402'): ('Panasonic',	'ERJ-2RKF3321X'),
    ('10KΩ 1%', 	'0402'): ('Panasonic',	'ERJ-PA2F1002X'),
    ('19.6KΩ 1%', 	'0402'): ('Yaego',	'AC0402FR-0719K6L'),
    ('30.9KΩ 1%', 	'0402'): ('Panasonic',	'ERJ-2RKF3092X'),
    ('45.3KΩ 1%',	'0402'): ('Panasonic',	'ERJ-2RKF4532X'),
    ('47KΩ 1%', 	'0402'): ('Panasonic',	'ERA-2AED473X'),
    ('68KΩ 1%', 	'0402'): ('Panasonic',	'ERA-2AED683X'),
    ('88.7KΩ 1%', 	'0402'): ('Yageo',	'AC0402FR-7D88K7L'),
    ('100KΩ 1%',	'0402'): ('Panasonic',	'ERJ-2RKF1003X'),
    ('143KΩ 1%',	'0402'): ('Panasonic',	'ERJ-2RKF1433X'),
    ('330KΩ 1%',	'0402'): ('Panasonic',	'ERJ-PA2F3303X'),
    ('464KΩ 1%',	'0402'): ('Panasonic',	'ERJ-2RKF4643X'),

    ('1KΩ@100MHz',	'0603'): ('Laird-Signal',	'MI0603J102R-10'),

    ('BLUE LED', '0603'): ('Rohm', 'SMLD12BN1WT86C'),
    ('RED LED', '0603'): ('Rohm', 'CSL0902UT1C'),
    ('YELLOW LED', '0603'): ('Rohm', 'CSL0901YT1C'),
    ('GREEN LED', '0603'): ('Rohm', 'CSL0902ET1C'),

    ('32.768kHz', 'Crystal_SMD_EuroQuartz_EQ161-2Pin_3.2x1.5mm'): ('Abracon', 'ABS07AIG-32.768KHZ-6-1-T'),

    ('NTCG103JF103FTDS 10KΩ@25C', '0402'): ('TDK', 'NTCG103JF103FTDS'),
}

other_components = {
    ('RB521CS30L,315', 'D_SOD-882'): None,
    ('LMK1C1106A', 'Texas_HTSSOP-14-1EP_4.4x5mm_P0.65mm_EP3.4x5mm_Mask3.155x3.255mm'): None,
    ('MAX31331TETB+', 'TDFN-10-1EP_3x3mm_P0.5mm_EP0.9x2mm'): None,
    ('AS3016204-0108X0PSAY', 'SOIC-8_5.3x5.3mm_P1.27mm'): None,
    ('BAV116WSQ-7', 'D_SOD-323F'): None,
    ('QPL9547', 'DFN-8-1EP_2x2mm_P0.5mm_EP0.8x1.6mm'): None,
    ('AX5043', 'QFN28'): None,
    ('TMS570LS0914APGEQQ1', 'TQFP-144_20x20mm_Pitch0.5mm'): None,
    ('2118714-2', 'TE_2118714-2'): None,
    ('SN3257QDYYRQ1', 'DYY0016A'): None,
    ('TCAN1044ADDFRQ1', 'TSOT-23-8'): None,
    ('FTSH-105-01-L-DV-K', 'FTSH-105-01-L-DV-K'): None,
    ('MAX4995AAUT+T', 'SOT-23-6_Handsoldering'): None,
    ('TQP7M9106', 'QFN24_TQP7M9104_QOR'): None,
    ('CONUFL001-SMD-T', 'CONN1_CONUFL_TEC'): None,
    ('MMCX-J-P-H-RA-TH1', 'COAX4_MMCX-J-P-H-RA-TH1_SAI'): None,
    ('2118718-2', 'TE_2118718-2'): None,
    ('TSW-103-08-F-S-RA', 'CON3_1X3_TR_TSW_SAI'): None,
    ('HTSW-102-07-G-S', 'CON2_1X2_TU_TSW'): None,
    ('STWD100NYWY3F', 'SOT-23-5'): None,
    ('AD4PS-1+', 'CJ725'): None,
    ('ESQ-126-39-G-D', 'CONN_ESQ-126-39-G-D_SAI'): None,
    ('SN74AHC1G02QDCKRQ1', 'DCK5'): None,
    ('74CBTLV1G125DBVRQ1', 'DBV5'): None,
    ('TPSM828302ARDSR', 'RDS0009A-MFG'): None,
    ('TPS62A02AQDRLRQ1', 'DRL0006A-MFG'): None,
    ('MP5073GG-P', 'QFN-12_MP5073_MNP'): None,
    ('O 16,0-JT22CT-A-P-3,3-LF', 'Oscillator_SMD_SiT_PQFN-4Pin_2.5x2.0mm'): None,
    ('DMP2037U-7', 'SOT-23'): None,
    ('ADL5501AKSZ-R7', 'KS-6_ADI'): None,
    ('SN74AHC1G08QDCKRQ1', 'DCK5'): None,
    ('SN74AHC1G09QDCKRQ1', 'DCK5'): None,
    ('BSS138', 'SOT-523'): None,
    ('MCP1799T-3302HTT', 'SOT-23'): None,
    ('QPC8010QTR7', 'QFN50P200X200X60-13N-D'): None,
    ('MPQ5072GG-AEC1', 'QFN-12_MP5073_MNP'): None,
    ('DAC5311IDCKRQ1', 'DCK6'): None,
    ('M0L1228QRGERQ1', 'VQFN24_4P1X4P1_TEX'): None,
    ('G125-MH11005L1P', 'G125-MH11005L1P'): None,
    ('', ''): None,
}

def xlat_value_to_partnum(s, footprint):
    pf = (s, footprint)
    if use_murata and pf in value_to_partnum_xlats_1b:
        v = value_to_partnum_xlats_1b[pf]
    elif pf in value_to_partnum_xlats_1:
        v = value_to_partnum_xlats_1[pf]
    elif do_stage2_xlat and pf in value_to_partnum_xlats_2:
        v = value_to_partnum_xlats_2[pf]
    else:
        if pf not in other_components:
            unknown_components.append("('" + s + "', '" + footprint + "'): ('', '')")
            pass
        v = ('', s)
        pass
    v = [v[0], v[1]]
    if value_in_comment:
        v[1] = v[1] + " " + s
        pass
    v[1] = v[1].replace(' ', ',')
    return v

if do_xls:
    # Output in Excel format
    wb = Workbook()
    ws = wb.active

    lineno = 1
    ws.cell(lineno, 1, 'Comment')
    ws.cell(lineno, 2, 'Designator')
    ws.cell(lineno, 3, 'Footprint')
    ws.cell(lineno, 4, 'Value')
    ws.cell(lineno, 5, 'Manufacturer')

    for line in cf:
        lineno += 1
        if len(line) != 8:
            sys.stderr.write("Line %s doesn't have 8 values, it has %d" %
                             (lineno, len(line)));
            sys.exit(1)
            pass
        designator = line[1]
        footprint = xlat_footprint(line[2]).strip('"')
        (mfg, partnum) = xlat_value_to_partnum(line[4], footprint)
        ws.cell(lineno, 1, partnum)
        ws.cell(lineno, 2, designator)
        ws.cell(lineno, 3, footprint)
        ws.cell(lineno, 4, line[4])
        ws.cell(lineno, 5, mfg)
        pass

    wb.save(outfn)
    pass
else:
    # Output in CSV format
    outfile = open(outfn, "w")
    ocf = csv.writer(outfile)
    lineno = 1
    ocf.writerow(('Comment', 'Designator', 'Footprint', 'Value',
                  'Manufacturer'))
    for line in cf:
        lineno += 1
        if len(line) != 8:
            sys.stderr.write("Line %s doesn't have 8 values, it has %d" %
                             (lineno, len(line)));
            sys.exit(1)
            pass
        designator = line[1]
        footprint = xlat_footprint(line[2]).strip('"')
        (mfg, partnum) = xlat_value_to_partnum(line[4], footprint)
        #comment = comment.replace('Ω', 'ohm')
        ocf.writerow((partnum, designator, footprint, line[4], mfg))
        pass
    pass

if unknown_components:
    print("Unknown components:")
    for i in unknown_components:
        print("  " + i)
        pass
    pass
